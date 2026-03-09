#!/usr/bin/env python3
"""Generate an inventory of Databricks assets via databricks-sdk-py.

Usage:
    python3 scripts/inventory_databricks_assets.py --root . --out databricks_as_is.md
        python3 scripts/inventory_databricks_assets.py --source sdk --out databricks_as_is.md

Environment (.env or env vars):
    DATABRICKS_HOST=https://<workspace-host>
    DATABRICKS_USERNAME=<admin-user>
    DATABRICKS_PASSWORD=<admin-password>
    # or use token auth
    DATABRICKS_TOKEN=<pat>
"""

from __future__ import annotations

import argparse
import json
import os
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from databricks.sdk import WorkspaceClient
from databricks.sdk.service import catalog, ml, sql, workspace


EXCLUDE_DIRS = {
    ".git",
    ".venv",
    "__pycache__",
    ".idea",
    ".vscode",
    "node_modules",
}

PATTERNS = [
    r"\\bdatabricks\\b",
    r"\\bdbutils\\b",
    r"\\bspark\\b",
    r"dbfs:/",
    r"\\bdelta\\b",
    r"\\bunity\\s*catalog\\b",
    r"\\bmlflow\\b",
    r"\\bjobs\\s*api\\b",
    r"\\bcluster[s]?\\b",
    r"\\bendpoints?\\b",
    r"\\bworkspace\\b",
]

PATTERN_RE = re.compile("|".join(PATTERNS), re.IGNORECASE)

SHEET_ORDER = [
    "Workspace Objects",
    "Jobs Runs",
    "Clusters Pools Policies",
    "SQL Warehouses",
    "SQL Dashboards Queries",
    "MLflow",
    "Unity Catalog",
    "External Locations",
    "Repos Git",
    "Secrets Tokens IP",
    "Identities",
    "Serving Vector Online",
    "Sharing",
    "DBFS",
]

KIND_TO_SHEET = {
    "workspace_dir": "Workspace Objects",
    "workspace_notebook": "Workspace Objects",
    "workspace_file": "Workspace Objects",
    "job": "Jobs Runs",
    "job_run": "Jobs Runs",
    "cluster": "Clusters Pools Policies",
    "instance_pool": "Clusters Pools Policies",
    "cluster_policy": "Clusters Pools Policies",
    "global_init_script": "Clusters Pools Policies",
    "instance_profile": "Clusters Pools Policies",
    "sql_warehouse": "SQL Warehouses",
    "sql_dashboard": "SQL Dashboards Queries",
    "lakeview_dashboard": "SQL Dashboards Queries",
    "sql_query": "SQL Dashboards Queries",
    "sql_alert": "SQL Dashboards Queries",
    "sql_alert_v2": "SQL Dashboards Queries",
    "mlflow_experiment": "MLflow",
    "mlflow_model": "MLflow",
    "model_version": "MLflow",
    "uc_catalog": "Unity Catalog",
    "uc_schema": "Unity Catalog",
    "uc_table": "Unity Catalog",
    "uc_volume": "Unity Catalog",
    "external_location": "External Locations",
    "storage_credential": "External Locations",
    "connection": "External Locations",
    "metastore": "External Locations",
    "repo": "Repos Git",
    "git_credential": "Repos Git",
    "secret_scope": "Secrets Tokens IP",
    "token": "Secrets Tokens IP",
    "ip_access_list": "Secrets Tokens IP",
    "user": "Identities",
    "group": "Identities",
    "service_principal": "Identities",
    "serving_endpoint": "Serving Vector Online",
    "vector_search_endpoint": "Serving Vector Online",
    "vector_search_index": "Serving Vector Online",
    "online_table": "Serving Vector Online",
    "share": "Sharing",
    "recipient": "Sharing",
    "provider": "Sharing",
    "dbfs_dir": "DBFS",
    "dbfs_file": "DBFS",
}


@dataclass
class Finding:
    path: str
    kind: str
    notes: str


def should_skip_dir(path: Path) -> bool:
    return path.name in EXCLUDE_DIRS


def iter_files(root: Path) -> Iterable[Path]:
    for dirpath, dirnames, filenames in os.walk(root):
        dirpath_p = Path(dirpath)
        dirnames[:] = [d for d in dirnames if not should_skip_dir(dirpath_p / d)]
        for name in filenames:
            yield dirpath_p / name


def detect_notebook(file_path: Path) -> Tuple[bool, str]:
    try:
        raw = file_path.read_text(encoding="utf-8")
    except Exception:
        return True, "notebook (unreadable)"
    try:
        data = json.loads(raw)
    except Exception:
        return True, "notebook (invalid json)"
    kernelspec = data.get("metadata", {}).get("kernelspec", {})
    lang = kernelspec.get("language") or "unknown"
    return True, f"notebook (kernel: {lang})"


def detect_databricks_reference(file_path: Path) -> Tuple[bool, str]:
    try:
        text = file_path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return False, ""
    if PATTERN_RE.search(text):
        return True, "reference found"
    return False, ""


def classify_file(file_path: Path) -> str:
    suffix = file_path.suffix.lower()
    if suffix == ".ipynb":
        return "notebook"
    if suffix in {".py", ".scala", ".sql", ".r"}:
        return "code"
    if suffix in {".json", ".yml", ".yaml", ".tf", ".tfvars", ".ini", ".toml"}:
        return "config"
    if suffix in {".md", ".txt"}:
        return "docs"
    if suffix in {".j2"}:
        return "template"
    return "other"


def collect_findings(root: Path) -> List[Finding]:
    findings: List[Finding] = []

    for path in iter_files(root):
        kind = classify_file(path)
        rel = str(path.relative_to(root))

        if kind == "notebook":
            _, note = detect_notebook(path)
            findings.append(Finding(rel, "notebook", note))
            continue

        if kind in {"code", "config", "docs", "template", "other"}:
            has_ref, note = detect_databricks_reference(path)
            if has_ref:
                findings.append(Finding(rel, kind, note))

    return findings


def load_env(env_path: Path) -> Dict[str, str]:
    if not env_path.exists():
        return {}
    env: Dict[str, str] = {}
    for line in env_path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip("'\"")
    return env


def build_workspace_client(root: Path) -> WorkspaceClient:
    env = load_env(root / ".env")
    host = env.get("DATABRICKS_HOST") or os.getenv("DATABRICKS_HOST", "")
    user = env.get("DATABRICKS_USER") or env.get("DATABRICKS_USERNAME") or os.getenv("DATABRICKS_USER", "")
    password = env.get("DATABRICKS_PASSWORD") or os.getenv("DATABRICKS_PASSWORD", "")
    token = env.get("DATABRICKS_TOKEN") or os.getenv("DATABRICKS_TOKEN", "")

    if host and not os.getenv("DATABRICKS_HOST"):
        os.environ["DATABRICKS_HOST"] = host
    if user and not os.getenv("DATABRICKS_USERNAME"):
        os.environ["DATABRICKS_USERNAME"] = user
    if password and not os.getenv("DATABRICKS_PASSWORD"):
        os.environ["DATABRICKS_PASSWORD"] = password
    if token and not os.getenv("DATABRICKS_TOKEN"):
        os.environ["DATABRICKS_TOKEN"] = token

    if not host:
        raise RuntimeError("Missing DATABRICKS_HOST. Set it in .env or environment.")

    if host and (user and password):
        return WorkspaceClient(host=host, username=user, password=password)
    if host and token:
        return WorkspaceClient(host=host, token=token)
    raise RuntimeError(
        "Missing Databricks credentials. Set DATABRICKS_USERNAME and DATABRICKS_PASSWORD or DATABRICKS_TOKEN."
    )


def safe_iter(label: str, iterator, warnings: List[str], batch_size: int, sleep_ms: int):
    try:
        count = 0
        for item in iterator:
            count += 1
            yield item
            if batch_size > 0 and count % batch_size == 0 and sleep_ms > 0:
                time.sleep(sleep_ms / 1000.0)
    except Exception as exc:
        warnings.append(f"{label} failed: {exc}")


def collect_findings_sdk(
    client: WorkspaceClient,
    include_runs: bool,
    include_query_history: bool,
    include_dbfs: bool,
    batch_size: int,
    batch_sleep_ms: int,
) -> Tuple[List[Finding], List[str]]:
    findings: List[Finding] = []
    warnings: List[str] = []

    # Workspace objects (notebooks, directories, files)
    stack = ["/"]
    while stack:
        current = stack.pop()
        for obj in safe_iter(
            "workspace.list",
            client.workspace.list(path=current),
            warnings,
            batch_size,
            batch_sleep_ms,
        ):
            obj_type = getattr(obj, "object_type", None)
            obj_path = getattr(obj, "path", "") or ""
            if str(obj_type) == "DIRECTORY":
                stack.append(obj_path)
                findings.append(Finding(obj_path, "workspace_dir", "workspace directory"))
            else:
                kind = "workspace_notebook" if str(obj_type) == "NOTEBOOK" else "workspace_file"
                lang = getattr(obj, "language", "unknown")
                findings.append(Finding(obj_path, kind, f"language: {lang}"))

    # Jobs
    for job in safe_iter("jobs.list", client.jobs.list(), warnings, batch_size, batch_sleep_ms):
        job_id = getattr(job, "job_id", None)
        name = getattr(getattr(job, "settings", None), "name", "")
        findings.append(Finding(f"job:{job_id}", "job", name))

    if include_runs and hasattr(client.jobs, "list_runs"):
        for run in safe_iter("jobs.list_runs", client.jobs.list_runs(), warnings, batch_size, batch_sleep_ms):
            run_id = getattr(run, "run_id", None)
            job_id = getattr(run, "job_id", None)
            state = getattr(getattr(run, "state", None), "life_cycle_state", "")
            findings.append(Finding(f"job-run:{run_id}", "job_run", f"job_id={job_id} state={state}"))

    # Clusters
    for cluster in safe_iter("clusters.list", client.clusters.list(), warnings, batch_size, batch_sleep_ms):
        cluster_id = getattr(cluster, "cluster_id", None)
        name = getattr(cluster, "cluster_name", "")
        findings.append(Finding(f"cluster:{cluster_id}", "cluster", name))

    for policy in safe_iter(
        "cluster_policies.list",
        client.cluster_policies.list(),
        warnings,
        batch_size,
        batch_sleep_ms,
    ):
        policy_id = getattr(policy, "policy_id", None)
        name = getattr(policy, "name", "")
        findings.append(Finding(f"cluster-policy:{policy_id}", "cluster_policy", name))

    if hasattr(client, "global_init_scripts"):
        for script in safe_iter(
            "global_init_scripts.list",
            client.global_init_scripts.list(),
            warnings,
            batch_size,
            batch_sleep_ms,
        ):
            script_id = getattr(script, "script_id", None)
            name = getattr(script, "name", "")
            findings.append(Finding(f"global-init-script:{script_id}", "global_init_script", name))

    if hasattr(client, "instance_profiles"):
        for prof in safe_iter(
            "instance_profiles.list",
            client.instance_profiles.list(),
            warnings,
            batch_size,
            batch_sleep_ms,
        ):
            arn = getattr(prof, "instance_profile_arn", "")
            findings.append(Finding(f"instance-profile:{arn}", "instance_profile", arn))

    # Instance Pools
    for pool in safe_iter("instance_pools.list", client.instance_pools.list(), warnings, batch_size, batch_sleep_ms):
        pool_id = getattr(pool, "instance_pool_id", None)
        name = getattr(pool, "instance_pool_name", "")
        findings.append(Finding(f"instance-pool:{pool_id}", "instance_pool", name))

    # SQL Warehouses
    for wh in safe_iter(
        "warehouses.list",
        client.warehouses.list(sql.ListWarehousesRequest(page_size=batch_size)),
        warnings,
        batch_size,
        batch_sleep_ms,
    ):
        wh_id = getattr(wh, "id", None)
        name = getattr(wh, "name", "")
        serverless = getattr(wh, "enable_serverless_compute", None)
        notes = f"{name} | serverless={serverless}"
        findings.append(Finding(f"sql-warehouse:{wh_id}", "sql_warehouse", notes))

    # Pipelines (Delta Live Tables)
    for pipeline in safe_iter(
        "pipelines.list",
        client.pipelines.list_pipelines(),
        warnings,
        batch_size,
        batch_sleep_ms,
    ):
        pipeline_id = getattr(pipeline, "pipeline_id", None)
        name = getattr(pipeline, "name", "")
        findings.append(Finding(f"pipeline:{pipeline_id}", "pipeline", name))

    # SQL assets (AI/BI dashboards, queries, alerts)
    for dash in safe_iter(
        "dashboards.list",
        client.dashboards.list(sql.ListDashboardsRequest(page_size=batch_size)),
        warnings,
        batch_size,
        batch_sleep_ms,
    ):
        dash_id = getattr(dash, "id", None)
        name = getattr(dash, "name", "")
        findings.append(Finding(f"sql-dashboard:{dash_id}", "sql_dashboard", name))

    for dash in safe_iter(
        "lakeview.list",
        client.lakeview.list(page_size=batch_size),
        warnings,
        batch_size,
        batch_sleep_ms,
    ):
        dash_id = getattr(dash, "dashboard_id", None) or getattr(dash, "id", None)
        name = getattr(dash, "display_name", "") or getattr(dash, "name", "")
        findings.append(Finding(f"lakeview-dashboard:{dash_id}", "lakeview_dashboard", name))

    for query in safe_iter(
        "queries.list",
        client.queries.list(sql.ListQueriesRequest(page_size=batch_size)),
        warnings,
        batch_size,
        batch_sleep_ms,
    ):
        query_id = getattr(query, "id", None)
        name = getattr(query, "name", "") or getattr(query, "display_name", "")
        findings.append(Finding(f"sql-query:{query_id}", "sql_query", name))

    if include_query_history and hasattr(client, "query_history"):
        for qh in safe_iter(
            "query_history.list",
            client.query_history.list(page_size=batch_size),
            warnings,
            batch_size,
            batch_sleep_ms,
        ):
            query_id = getattr(qh, "query_id", None)
            status = getattr(qh, "status", None)
            findings.append(Finding(f"query-history:{query_id}", "query_history", str(status)))

    for alert in safe_iter(
        "alerts.list",
        client.alerts.list(sql.ListAlertsRequest(page_size=batch_size)),
        warnings,
        batch_size,
        batch_sleep_ms,
    ):
        alert_id = getattr(alert, "id", None)
        name = getattr(alert, "name", "") or getattr(alert, "display_name", "")
        findings.append(Finding(f"sql-alert:{alert_id}", "sql_alert", name))

    for alert in safe_iter(
        "alerts_v2.list",
        client.alerts_v2.list_alerts(page_size=batch_size),
        warnings,
        batch_size,
        batch_sleep_ms,
    ):
        alert_id = getattr(alert, "id", None)
        name = getattr(alert, "display_name", "")
        findings.append(Finding(f"sql-alert-v2:{alert_id}", "sql_alert_v2", name))

    # MLflow assets
    for exp in safe_iter(
        "experiments.list",
        client.experiments.list_experiments(ml.ListExperimentsRequest(max_results=batch_size)),
        warnings,
        batch_size,
        batch_sleep_ms,
    ):
        exp_id = getattr(exp, "experiment_id", None)
        name = getattr(exp, "name", "")
        findings.append(Finding(f"mlflow-experiment:{exp_id}", "mlflow_experiment", name))

    for model in safe_iter(
        "registered_models.list",
        client.registered_models.list(),
        warnings,
        batch_size,
        batch_sleep_ms,
    ):
        name = getattr(model, "name", "")
        findings.append(Finding(f"mlflow-model:{name}", "mlflow_model", name))

    if hasattr(client, "model_versions"):
        for mv in safe_iter(
            "model_versions.list",
            client.model_versions.list(),
            warnings,
            batch_size,
            batch_sleep_ms,
        ):
            name = getattr(mv, "name", "")
            version = getattr(mv, "version", "")
            findings.append(Finding(f"model-version:{name}:{version}", "model_version", name))

    if hasattr(client, "feature_store"):
        for fs in safe_iter(
            "feature_store.list",
            client.feature_store.list(),
            warnings,
            batch_size,
            batch_sleep_ms,
        ):
            name = getattr(fs, "name", "")
            findings.append(Finding(f"feature-store:{name}", "feature_store", name))

    if hasattr(client, "feature_engineering"):
        for fe in safe_iter(
            "feature_engineering.list",
            client.feature_engineering.list(),
            warnings,
            batch_size,
            batch_sleep_ms,
        ):
            name = getattr(fe, "name", "")
            findings.append(Finding(f"feature-engineering:{name}", "feature_engineering", name))

    # Unity Catalog (catalogs, schemas, tables, volumes)
    for catalog_obj in safe_iter(
        "catalogs.list",
        client.catalogs.list(catalog.ListCatalogsRequest(max_results=batch_size)),
        warnings,
        batch_size,
        batch_sleep_ms,
    ):
        catalog_name = getattr(catalog_obj, "name", "")
        findings.append(Finding(f"uc-catalog:{catalog_name}", "uc_catalog", catalog_name))

        for schema in safe_iter(
            f"schemas.list({catalog_name})",
            client.schemas.list(catalog.ListSchemasRequest(catalog_name=catalog_name)),
            warnings,
            batch_size,
            batch_sleep_ms,
        ):
            schema_name = getattr(schema, "name", "")
            findings.append(Finding(f"uc-schema:{catalog_name}.{schema_name}", "uc_schema", schema_name))

            for table in safe_iter(
                f"tables.list({catalog_name}.{schema_name})",
                client.tables.list(
                    catalog.ListTablesRequest(catalog_name=catalog_name, schema_name=schema_name)
                ),
                warnings,
                batch_size,
                batch_sleep_ms,
            ):
                table_name = getattr(table, "name", "")
                findings.append(
                    Finding(
                        f"uc-table:{catalog_name}.{schema_name}.{table_name}",
                        "uc_table",
                        getattr(table, "table_type", ""),
                    )
                )

            for vol in safe_iter(
                f"volumes.list({catalog_name}.{schema_name})",
                client.volumes.list(catalog_name=catalog_name, schema_name=schema_name),
                warnings,
                batch_size,
                batch_sleep_ms,
            ):
                vol_name = getattr(vol, "name", "")
                findings.append(Finding(f"uc-volume:{vol_name}", "uc_volume", vol_name))

    for eloc in safe_iter(
        "external_locations.list",
        client.external_locations.list(),
        warnings,
        batch_size,
        batch_sleep_ms,
    ):
        name = getattr(eloc, "name", "")
        url = getattr(eloc, "url", "")
        findings.append(Finding(f"external-location:{name}", "external_location", url))

    for cred in safe_iter(
        "storage_credentials.list",
        client.storage_credentials.list(),
        warnings,
        batch_size,
        batch_sleep_ms,
    ):
        name = getattr(cred, "name", "")
        findings.append(Finding(f"storage-credential:{name}", "storage_credential", name))

    for conn in safe_iter(
        "connections.list",
        client.connections.list(),
        warnings,
        batch_size,
        batch_sleep_ms,
    ):
        name = getattr(conn, "name", "")
        type_name = getattr(conn, "connection_type", "")
        findings.append(Finding(f"connection:{name}", "connection", str(type_name)))

    if hasattr(client, "metastores"):
        for ms in safe_iter(
            "metastores.list",
            client.metastores.list(),
            warnings,
            batch_size,
            batch_sleep_ms,
        ):
            name = getattr(ms, "name", "")
            ms_id = getattr(ms, "metastore_id", "")
            findings.append(Finding(f"metastore:{ms_id}", "metastore", name))

    # Repos
    for repo in safe_iter(
        "repos.list",
        client.repos.list(workspace.ListReposRequest()),
        warnings,
        batch_size,
        batch_sleep_ms,
    ):
        repo_id = getattr(repo, "id", None)
        path = getattr(repo, "path", "")
        findings.append(Finding(f"repo:{repo_id}", "repo", path))

    if hasattr(client, "git_credentials"):
        for cred in safe_iter(
            "git_credentials.list",
            client.git_credentials.list(),
            warnings,
            batch_size,
            batch_sleep_ms,
        ):
            cred_id = getattr(cred, "credential_id", None)
            user = getattr(cred, "git_username", "")
            findings.append(Finding(f"git-credential:{cred_id}", "git_credential", user))

    # Secrets scopes
    for scope in safe_iter(
        "secrets.list_scopes",
        client.secrets.list_scopes(),
        warnings,
        batch_size,
        batch_sleep_ms,
    ):
        name = getattr(scope, "name", "")
        findings.append(Finding(f"secret-scope:{name}", "secret_scope", name))

    # Tokens (admin only)
    for token in safe_iter(
        "tokens.list",
        client.tokens.list(),
        warnings,
        batch_size,
        batch_sleep_ms,
    ):
        token_id = getattr(token, "token_id", None)
        creator = getattr(token, "created_by_username", "")
        findings.append(Finding(f"token:{token_id}", "token", f"created_by={creator}"))

    if hasattr(client, "ip_access_lists"):
        for ip in safe_iter(
            "ip_access_lists.list",
            client.ip_access_lists.list(),
            warnings,
            batch_size,
            batch_sleep_ms,
        ):
            list_id = getattr(ip, "list_id", "")
            label = getattr(ip, "label", "")
            findings.append(Finding(f"ip-access-list:{list_id}", "ip_access_list", label))

    # SCIM identities
    for user_obj in safe_iter("users.list", client.users.list(), warnings, batch_size, batch_sleep_ms):
        name = getattr(user_obj, "user_name", "") or getattr(user_obj, "userName", "")
        user_id = getattr(user_obj, "id", "")
        findings.append(Finding(f"user:{user_id}", "user", name))

    for group in safe_iter("groups.list", client.groups.list(), warnings, batch_size, batch_sleep_ms):
        name = getattr(group, "display_name", "") or getattr(group, "displayName", "")
        group_id = getattr(group, "id", "")
        findings.append(Finding(f"group:{group_id}", "group", name))

    for sp in safe_iter(
        "service_principals.list",
        client.service_principals.list(),
        warnings,
        batch_size,
        batch_sleep_ms,
    ):
        name = getattr(sp, "display_name", "") or getattr(sp, "displayName", "")
        sp_id = getattr(sp, "id", "")
        findings.append(Finding(f"service-principal:{sp_id}", "service_principal", name))

    for endpoint in safe_iter(
        "serving_endpoints.list",
        client.serving_endpoints.list(),
        warnings,
        batch_size,
        batch_sleep_ms,
    ):
        name = getattr(endpoint, "name", "")
        findings.append(Finding(f"serving-endpoint:{name}", "serving_endpoint", name))

    if hasattr(client, "vector_search_endpoints"):
        for vse in safe_iter(
            "vector_search_endpoints.list",
            client.vector_search_endpoints.list(),
            warnings,
            batch_size,
            batch_sleep_ms,
        ):
            name = getattr(vse, "name", "")
            findings.append(Finding(f"vector-search-endpoint:{name}", "vector_search_endpoint", name))

    if hasattr(client, "vector_search_indexes"):
        for vsi in safe_iter(
            "vector_search_indexes.list",
            client.vector_search_indexes.list(),
            warnings,
            batch_size,
            batch_sleep_ms,
        ):
            name = getattr(vsi, "name", "")
            findings.append(Finding(f"vector-search-index:{name}", "vector_search_index", name))

    if hasattr(client, "online_tables"):
        for ot in safe_iter(
            "online_tables.list",
            client.online_tables.list(),
            warnings,
            batch_size,
            batch_sleep_ms,
        ):
            name = getattr(ot, "name", "")
            findings.append(Finding(f"online-table:{name}", "online_table", name))

    if hasattr(client, "shares"):
        for share in safe_iter(
            "shares.list",
            client.shares.list(),
            warnings,
            batch_size,
            batch_sleep_ms,
        ):
            name = getattr(share, "name", "")
            findings.append(Finding(f"share:{name}", "share", name))

    if hasattr(client, "recipients"):
        for rec in safe_iter(
            "recipients.list",
            client.recipients.list(),
            warnings,
            batch_size,
            batch_sleep_ms,
        ):
            name = getattr(rec, "name", "")
            findings.append(Finding(f"recipient:{name}", "recipient", name))

    if hasattr(client, "providers"):
        for prov in safe_iter(
            "providers.list",
            client.providers.list(),
            warnings,
            batch_size,
            batch_sleep_ms,
        ):
            name = getattr(prov, "name", "")
            findings.append(Finding(f"provider:{name}", "provider", name))

    if include_dbfs and hasattr(client, "dbfs"):
        for file_info in safe_iter(
            "dbfs.list",
            client.dbfs.list(path="/"),
            warnings,
            batch_size,
            batch_sleep_ms,
        ):
            path = getattr(file_info, "path", "")
            is_dir = getattr(file_info, "is_dir", False)
            kind = "dbfs_dir" if is_dir else "dbfs_file"
            findings.append(Finding(path, kind, "dbfs root"))

    return findings, warnings


def summarize(findings: List[Finding], warnings: List[str]) -> Tuple[str, List[str]]:
    counts = {}
    for f in findings:
        counts[f.kind] = counts.get(f.kind, 0) + 1

    summary_lines = ["# Databricks AS-IS Inventory", ""]
    summary_lines.append("## Summary")
    summary_lines.append("")
    for kind in sorted(counts.keys()):
        summary_lines.append(f"- {kind}: {counts[kind]}")
    summary_lines.append("")

    detail_lines = ["## Findings", ""]
    for f in sorted(findings, key=lambda x: (x.kind, x.path)):
        detail_lines.append(f"- [{f.kind}] {f.path} ({f.notes})")

    if warnings:
        detail_lines.append("")
        detail_lines.append("## Warnings")
        detail_lines.append("")
        for warning in warnings:
            detail_lines.append(f"- {warning}")

    return "\n".join(summary_lines), detail_lines


def write_excel(findings: List[Finding], warnings: List[str], out_path: Path) -> None:
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError("Missing dependency: openpyxl. Install with `pip install openpyxl`.") from exc

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    sheets = {}
    for sheet_name in SHEET_ORDER:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["kind", "path", "notes"])
        sheets[sheet_name] = ws

    for item in findings:
        sheet_name = KIND_TO_SHEET.get(item.kind, "Workspace Objects")
        sheets[sheet_name].append([item.kind, item.path, item.notes])

    if warnings:
        ws = wb.create_sheet(title="Warnings")
        ws.append(["warning"])
        for warning in warnings:
            ws.append([warning])

    wb.save(out_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Inventory Databricks assets in a workspace.")
    parser.add_argument("--root", default=".", help="Workspace root to scan")
    parser.add_argument("--out", default="databricks_as_is.md", help="Output markdown file")
    parser.add_argument("--out-xlsx", default="", help="Output Excel file with categorized sheets")
    parser.add_argument(
        "--source",
        default="sdk",
        choices=["sdk", "local"],
        help="Inventory source: sdk (databricks-sdk-py) or local (workspace scan)",
    )
    parser.add_argument(
        "--include-runs",
        action="store_true",
        help="Include job runs (can be large)",
    )
    parser.add_argument(
        "--include-query-history",
        action="store_true",
        help="Include SQL query history (can be large)",
    )
    parser.add_argument(
        "--include-dbfs",
        action="store_true",
        help="Include DBFS root listing",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=200,
        help="Number of items per batch before sleeping",
    )
    parser.add_argument(
        "--batch-sleep-ms",
        type=int,
        default=0,
        help="Sleep time in ms between batches",
    )
    args = parser.parse_args()

    root = Path(args.root).resolve()

    if args.source == "sdk":
        client = build_workspace_client(root)
        findings, warnings = collect_findings_sdk(
            client,
            include_runs=args.include_runs,
            include_query_history=args.include_query_history,
            include_dbfs=args.include_dbfs,
            batch_size=args.batch_size,
            batch_sleep_ms=args.batch_sleep_ms,
        )
    else:
        findings = collect_findings(root)
        warnings = []
    summary, details = summarize(findings, warnings)

    out_path = Path(args.out)
    if not out_path.is_absolute():
        out_path = root / out_path

    out_path.write_text(summary + "\n" + "\n".join(details) + "\n", encoding="utf-8")
    if args.out_xlsx:
        xlsx_path = Path(args.out_xlsx)
        if not xlsx_path.is_absolute():
            xlsx_path = root / xlsx_path
        write_excel(findings, warnings, xlsx_path)
    print(f"Wrote {len(findings)} findings to {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
