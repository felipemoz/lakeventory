from pathlib import Path

from lakeventory.models import Finding
from lakeventory.output import summarize, write_markdown, write_excel


def test_summarize_includes_warnings():
    findings = [Finding("/a", "kind_a", "notes")]
    warnings = ["warn"]

    summary, details = summarize(findings, warnings)

    assert "kind_a" in summary
    assert any("## Warnings" in line for line in details)


def test_write_markdown(tmp_path: Path):
    findings = [Finding("/a", "kind_a", "notes", lockin_count=2, lockin_details="aws:2")]
    warnings = []
    out_path = tmp_path / "out.md"

    write_markdown(findings, warnings, out_path)

    text = out_path.read_text(encoding="utf-8")
    assert "Databricks AS-IS Inventory" in text
    assert "kind_a" in text
    assert "lockin_count=2" in text


def test_write_excel(tmp_path: Path):
    findings = [Finding("/a", "workspace_file", "notes", lockin_count=1, lockin_details="azure:1")]
    warnings = ["warn"]
    out_path = tmp_path / "out.xlsx"

    write_excel(findings, warnings, out_path)

    assert out_path.exists()
    try:
        from openpyxl import load_workbook
    except Exception:
        return

    wb = load_workbook(out_path)
    assert "Workspace Objects" in wb.sheetnames
    assert "Warnings" in wb.sheetnames
    ws = wb["Workspace Objects"]
    headers = [cell.value for cell in ws[1]]
    assert "lockin_count" in headers
    assert "lockin_details" in headers
