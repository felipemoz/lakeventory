"""Tests for multi-workspace orchestration."""

from pathlib import Path

import pytest

from lakeventory.models import Finding
from lakeventory.multi_workspace import (
    ComparisonConfig,
    MultiWorkspaceConfig,
    WorkspaceConfig,
    build_workspace_client_from_env_file,
    load_workspaces_config,
    write_comparison_report,
)


# ---------------------------------------------------------------------------
# load_workspaces_config
# ---------------------------------------------------------------------------


def test_load_workspaces_config_minimal(tmp_path: Path):
    """Parse a minimal valid workspaces.yaml."""
    config_file = tmp_path / "workspaces.yaml"
    config_file.write_text(
        """
workspaces:
  - name: dev
    env_file: .env.dev
  - name: prod
    env_file: .env.prod
    output_dir: .reports/prod
    batch_size: 100
    serverless: true
""",
        encoding="utf-8",
    )

    config = load_workspaces_config(config_file)

    assert len(config.workspaces) == 2

    dev = config.workspaces[0]
    assert dev.name == "dev"
    assert dev.env_file == ".env.dev"
    assert dev.batch_size == 200          # default
    assert dev.serverless is False        # default
    assert dev.collectors == ""           # default

    prod = config.workspaces[1]
    assert prod.name == "prod"
    assert prod.output_dir == ".reports/prod"
    assert prod.batch_size == 100
    assert prod.serverless is True


def test_load_workspaces_config_comparison_defaults(tmp_path: Path):
    """Comparison section defaults are applied when omitted."""
    config_file = tmp_path / "workspaces.yaml"
    config_file.write_text(
        """
workspaces:
  - name: dev
    env_file: .env.dev
""",
        encoding="utf-8",
    )

    config = load_workspaces_config(config_file)

    assert config.comparison.output_dir == ".reports"
    assert config.comparison.out_xlsx == "compare_workspaces.xlsx"


def test_load_workspaces_config_comparison_custom(tmp_path: Path):
    """Custom comparison settings are parsed correctly."""
    config_file = tmp_path / "workspaces.yaml"
    config_file.write_text(
        """
workspaces:
  - name: dev
    env_file: .env.dev

comparison:
  output_dir: .custom/reports
  out_xlsx: compare_dev_vs_prod.xlsx
""",
        encoding="utf-8",
    )

    config = load_workspaces_config(config_file)

    assert config.comparison.output_dir == ".custom/reports"
    assert config.comparison.out_xlsx == "compare_dev_vs_prod.xlsx"


def test_load_workspaces_config_all_fields(tmp_path: Path):
    """All optional fields are parsed."""
    config_file = tmp_path / "workspaces.yaml"
    config_file.write_text(
        """
workspaces:
  - name: staging
    env_file: .env.staging
    output_dir: .reports/staging
    batch_size: 50
    batch_sleep_ms: 100
    serverless: false
    collectors: workspace,jobs
    include_runs: true
    include_query_history: true
    include_dbfs: true
""",
        encoding="utf-8",
    )

    config = load_workspaces_config(config_file)

    ws = config.workspaces[0]
    assert ws.batch_sleep_ms == 100
    assert ws.collectors == "workspace,jobs"
    assert ws.include_runs is True
    assert ws.include_query_history is True
    assert ws.include_dbfs is True


def test_load_workspaces_config_file_not_found(tmp_path: Path):
    """FileNotFoundError is raised when config file is missing."""
    with pytest.raises(FileNotFoundError, match="not found"):
        load_workspaces_config(tmp_path / "missing.yaml")


def test_load_workspaces_config_missing_name(tmp_path: Path):
    """ValueError is raised when workspace entry is missing 'name'."""
    config_file = tmp_path / "workspaces.yaml"
    config_file.write_text(
        """
workspaces:
  - env_file: .env.dev
""",
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="'name' field"):
        load_workspaces_config(config_file)


def test_load_workspaces_config_missing_env_file(tmp_path: Path):
    """ValueError is raised when workspace entry is missing 'env_file'."""
    config_file = tmp_path / "workspaces.yaml"
    config_file.write_text(
        """
workspaces:
  - name: dev
""",
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="'env_file' field"):
        load_workspaces_config(config_file)


def test_load_workspaces_config_invalid_yaml_type(tmp_path: Path):
    """ValueError is raised when YAML root is not a mapping."""
    config_file = tmp_path / "workspaces.yaml"
    config_file.write_text("- item1\n- item2\n", encoding="utf-8")

    with pytest.raises(ValueError, match="YAML mapping"):
        load_workspaces_config(config_file)


def test_load_workspaces_config_empty_workspaces(tmp_path: Path):
    """Empty workspace list is handled gracefully."""
    config_file = tmp_path / "workspaces.yaml"
    config_file.write_text("workspaces: []\n", encoding="utf-8")

    config = load_workspaces_config(config_file)
    assert config.workspaces == []


# ---------------------------------------------------------------------------
# build_workspace_client_from_env_file
# ---------------------------------------------------------------------------


def test_build_workspace_client_from_env_file_token(monkeypatch, tmp_path: Path):
    """Builds a client using a token from an env file."""
    import lakeventory.multi_workspace as mw

    class FakeWorkspaceClient:
        def __init__(self, **kwargs):
            self.kwargs = kwargs

    monkeypatch.setattr(mw, "WorkspaceClient", FakeWorkspaceClient, raising=False)

    # Patch inside the function (it imports locally)
    import databricks.sdk as sdk_module

    monkeypatch.setattr(sdk_module, "WorkspaceClient", FakeWorkspaceClient)

    env_file = tmp_path / ".env.dev"
    env_file.write_text(
        "DATABRICKS_HOST=https://adb-1234567890.1.azuredatabricks.net\nDATABRICKS_TOKEN=mytoken\n",
        encoding="utf-8",
    )

    client = build_workspace_client_from_env_file(env_file)

    assert isinstance(client, FakeWorkspaceClient)
    assert client.kwargs["host"] == "https://adb-1234567890.1.azuredatabricks.net"
    assert client.kwargs["token"] == "mytoken"


def test_build_workspace_client_from_env_file_missing_host(tmp_path: Path):
    """RuntimeError is raised when DATABRICKS_HOST is missing."""
    env_file = tmp_path / ".env.dev"
    env_file.write_text("DATABRICKS_TOKEN=mytoken\n", encoding="utf-8")

    with pytest.raises(RuntimeError, match="Missing DATABRICKS_HOST"):
        build_workspace_client_from_env_file(env_file)


def test_build_workspace_client_from_env_file_no_credentials(tmp_path: Path):
    """RuntimeError is raised when no credentials are configured."""
    env_file = tmp_path / ".env.dev"
    env_file.write_text(
        "DATABRICKS_HOST=https://adb-123.azuredatabricks.net\n", encoding="utf-8"
    )

    with pytest.raises(RuntimeError, match="Missing Databricks credentials"):
        build_workspace_client_from_env_file(env_file)


# ---------------------------------------------------------------------------
# write_comparison_report
# ---------------------------------------------------------------------------


def test_write_comparison_report_creates_file(tmp_path: Path):
    """Comparison report Excel file is created."""
    findings_dev = [
        Finding("/notebooks/a", "workspace_notebook", "nb a"),
        Finding("/jobs/1", "job", "job 1"),
    ]
    findings_prod = [
        Finding("/notebooks/b", "workspace_notebook", "nb b"),
        Finding("/notebooks/c", "workspace_notebook", "nb c"),
        Finding("/clusters/1", "cluster", "cluster 1"),
    ]

    out_path = tmp_path / "compare.xlsx"
    write_comparison_report({"dev": findings_dev, "prod": findings_prod}, out_path)

    assert out_path.exists()


def test_write_comparison_report_summary_sheet(tmp_path: Path):
    """Summary sheet contains workspace names as columns and asset type rows."""
    from openpyxl import load_workbook

    findings_dev = [
        Finding("/notebooks/a", "workspace_notebook", "nb a"),
        Finding("/notebooks/b", "workspace_notebook", "nb b"),
        Finding("/jobs/1", "job", "job 1"),
    ]
    findings_prod = [
        Finding("/notebooks/c", "workspace_notebook", "nb c"),
        Finding("/clusters/1", "cluster", "cluster 1"),
    ]

    out_path = tmp_path / "compare.xlsx"
    write_comparison_report({"dev": findings_dev, "prod": findings_prod}, out_path)

    wb = load_workbook(out_path)
    assert "Summary" in wb.sheetnames

    ws = wb["Summary"]
    headers = [cell.value for cell in ws[1]]
    assert headers[0] == "Asset Type"
    assert "dev" in headers
    assert "prod" in headers

    # Check that asset types appear as rows
    row_values = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
    asset_types = [row[0] for row in row_values]
    assert "workspace_notebook" in asset_types
    assert "job" in asset_types
    assert "cluster" in asset_types

    # Verify counts for workspace_notebook
    dev_col = headers.index("dev")
    prod_col = headers.index("prod")
    nb_row = next(r for r in row_values if r[0] == "workspace_notebook")
    assert nb_row[dev_col] == 2
    assert nb_row[prod_col] == 1


def test_write_comparison_report_per_workspace_sheets(tmp_path: Path):
    """Each workspace has its own detail sheet."""
    from openpyxl import load_workbook

    findings_dev = [Finding("/notebooks/a", "workspace_notebook", "nb")]
    findings_prod = [Finding("/jobs/1", "job", "job")]

    out_path = tmp_path / "compare.xlsx"
    write_comparison_report({"dev": findings_dev, "prod": findings_prod}, out_path)

    wb = load_workbook(out_path)
    assert "dev" in wb.sheetnames
    assert "prod" in wb.sheetnames

    ws_dev = wb["dev"]
    headers = [cell.value for cell in ws_dev[1]]
    assert "kind" in headers
    assert "path" in headers


def test_write_comparison_report_total_row(tmp_path: Path):
    """TOTAL row is present in Summary sheet."""
    from openpyxl import load_workbook

    findings = [Finding("/a", "job", "j"), Finding("/b", "job", "j2")]
    out_path = tmp_path / "compare.xlsx"
    write_comparison_report({"ws1": findings}, out_path)

    wb = load_workbook(out_path)
    ws = wb["Summary"]
    last_row = [cell.value for cell in list(ws.iter_rows())[-1]]
    assert last_row[0] == "TOTAL"
    assert last_row[1] == 2


def test_write_comparison_report_empty_workspace(tmp_path: Path):
    """Works correctly when a workspace has no findings."""
    out_path = tmp_path / "compare.xlsx"
    write_comparison_report({"empty_ws": []}, out_path)

    assert out_path.exists()
