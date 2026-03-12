"""Multi-workspace inventory orchestration."""

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from lakeventory.collectors import collect_all_findings, collect_findings_selective
from lakeventory.models import Finding

logger = logging.getLogger(__name__)


def _parse_env_file(env_file: Path) -> Dict[str, str]:
    """Parse environment variables from a specific .env file (legacy multi-workspace support)."""
    env: Dict[str, str] = {}
    if not env_file.exists():
        return env
    for line in env_file.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


@dataclass
class WorkspaceConfig:
    """Configuration for a single workspace inventory run."""

    name: str
    env_file: str
    output_dir: str = ".reports"
    batch_size: int = 200
    batch_sleep_ms: int = 0
    serverless: bool = False
    collectors: str = ""
    include_runs: bool = False
    include_query_history: bool = False
    include_dbfs: bool = False


@dataclass
class ComparisonConfig:
    """Configuration for the consolidated comparison report."""

    output_dir: str = ".reports"
    out_xlsx: str = "compare_workspaces.xlsx"


@dataclass
class MultiWorkspaceConfig:
    """Full multi-workspace configuration."""

    workspaces: List[WorkspaceConfig] = field(default_factory=list)
    comparison: ComparisonConfig = field(default_factory=ComparisonConfig)


def load_workspaces_config(config_path: Path) -> MultiWorkspaceConfig:
    """Parse a workspaces.yaml configuration file.

    Args:
        config_path: Path to the workspaces YAML configuration file

    Returns:
        MultiWorkspaceConfig with workspace and comparison settings

    Raises:
        FileNotFoundError: If config_path does not exist
        RuntimeError: If pyyaml is not installed
        ValueError: If config format is invalid
    """
    try:
        import yaml
    except ImportError as exc:
        raise RuntimeError(
            "Missing dependency: pyyaml. Install with `pip install pyyaml`."
        ) from exc

    if not config_path.exists():
        raise FileNotFoundError(f"Workspaces config not found: {config_path}")

    raw = yaml.safe_load(config_path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise ValueError("workspaces.yaml must contain a YAML mapping at the top level")

    workspaces = []
    for entry in raw.get("workspaces", []):
        if "name" not in entry:
            raise ValueError("Each workspace entry must have a 'name' field")
        if "env_file" not in entry:
            raise ValueError(f"Workspace '{entry['name']}' must have an 'env_file' field")
        workspaces.append(
            WorkspaceConfig(
                name=entry["name"],
                env_file=entry["env_file"],
                output_dir=entry.get("output_dir", ".reports"),
                batch_size=int(entry.get("batch_size", 200)),
                batch_sleep_ms=int(entry.get("batch_sleep_ms", 0)),
                serverless=bool(entry.get("serverless", False)),
                collectors=entry.get("collectors", "") or "",
                include_runs=bool(entry.get("include_runs", False)),
                include_query_history=bool(entry.get("include_query_history", False)),
                include_dbfs=bool(entry.get("include_dbfs", False)),
            )
        )

    comp_raw = raw.get("comparison", {}) or {}
    comparison = ComparisonConfig(
        output_dir=comp_raw.get("output_dir", ".reports"),
        out_xlsx=comp_raw.get("out_xlsx", "compare_workspaces.xlsx"),
    )

    return MultiWorkspaceConfig(workspaces=workspaces, comparison=comparison)


def build_workspace_client_from_env_file(env_file: Path):
    """Build WorkspaceClient using credentials from a specific env file.

    Unlike build_workspace_client(), this function does not modify os.environ,
    making it safe to use in sequential multi-workspace runs.

    Args:
        env_file: Path to the .env file for this workspace

    Returns:
        Configured WorkspaceClient instance

    Raises:
        RuntimeError: If required credentials are missing
    """
    from databricks.sdk import WorkspaceClient

    env = _parse_env_file(env_file)

    host = env.get("DATABRICKS_HOST", "")
    if not host:
        raise RuntimeError(f"Missing DATABRICKS_HOST in {env_file}")

    client_id = env.get("DATABRICKS_CLIENT_ID", "")
    client_secret = env.get("DATABRICKS_CLIENT_SECRET", "")
    token = env.get("DATABRICKS_TOKEN", "")
    user = env.get("DATABRICKS_USERNAME") or env.get("DATABRICKS_USER", "")
    password = env.get("DATABRICKS_PASSWORD", "")

    if client_id and client_secret:
        return WorkspaceClient(host=host, client_id=client_id, client_secret=client_secret)
    if token:
        return WorkspaceClient(host=host, token=token)
    if user and password:
        return WorkspaceClient(host=host, username=user, password=password)

    raise RuntimeError(
        f"Missing Databricks credentials in {env_file}. Configure one of:\n"
        "  1. Service Principal: DATABRICKS_CLIENT_ID + DATABRICKS_CLIENT_SECRET\n"
        "  2. PAT Token: DATABRICKS_TOKEN\n"
        "  3. Basic Auth: DATABRICKS_USERNAME + DATABRICKS_PASSWORD"
    )


def run_workspace_inventory(
    ws_config: WorkspaceConfig,
    base_dir: Path,
) -> Tuple[str, List[Finding], List[str]]:
    """Run inventory for a single workspace.

    Args:
        ws_config: Workspace configuration
        base_dir: Base directory for resolving relative env_file paths

    Returns:
        Tuple of (workspace_id, findings, warnings)
    """
    from lakeventory.inventory_cli import _extract_workspace_id

    env_file = (base_dir / ws_config.env_file).resolve()
    logger.info("Running inventory for workspace '%s' (%s)", ws_config.name, env_file)

    client = build_workspace_client_from_env_file(env_file)

    env = _parse_env_file(env_file)
    workspace_id = _extract_workspace_id(env.get("DATABRICKS_HOST", "")) or ws_config.name

    serverless_collectors = (
        "workspace,jobs,sql,mlflow,unity_catalog,repos,security,identities,serving,sharing,dbfs"
    )

    if ws_config.serverless and not ws_config.collectors:
        findings, warnings = collect_findings_selective(
            client,
            collectors=serverless_collectors,
            include_runs=ws_config.include_runs,
            include_query_history=ws_config.include_query_history,
            batch_size=ws_config.batch_size,
            batch_sleep_ms=ws_config.batch_sleep_ms,
        )
    elif ws_config.collectors:
        findings, warnings = collect_findings_selective(
            client,
            collectors=ws_config.collectors,
            include_runs=ws_config.include_runs,
            include_query_history=ws_config.include_query_history,
            batch_size=ws_config.batch_size,
            batch_sleep_ms=ws_config.batch_sleep_ms,
        )
    else:
        findings, warnings = collect_all_findings(
            client,
            include_runs=ws_config.include_runs,
            include_query_history=ws_config.include_query_history,
            include_dbfs=ws_config.include_dbfs,
            batch_size=ws_config.batch_size,
            batch_sleep_ms=ws_config.batch_sleep_ms,
        )

    return workspace_id, findings, warnings


def write_comparison_report(
    workspace_results: Dict[str, List[Finding]],
    out_path: Path,
) -> None:
    """Write a consolidated comparison Excel report across multiple workspaces.

    The report contains:
    - A Summary sheet with asset-type counts per workspace side by side
    - One detail sheet per workspace listing all its findings

    Args:
        workspace_results: Mapping of workspace name -> list of findings
        out_path: Path to output Excel file

    Raises:
        RuntimeError: If openpyxl is not installed
    """
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError(
            "Missing dependency: openpyxl. Install with `pip install openpyxl`."
        ) from exc

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    workspace_names = list(workspace_results.keys())

    # Compute per-workspace kind counts
    all_kinds: List[str] = []
    counts_per_ws: Dict[str, Dict[str, int]] = {}
    for ws_name, findings in workspace_results.items():
        counts: Dict[str, int] = {}
        for f in findings:
            counts[f.kind] = counts.get(f.kind, 0) + 1
            if f.kind not in all_kinds:
                all_kinds.append(f.kind)
        counts_per_ws[ws_name] = counts

    all_kinds_sorted = sorted(all_kinds)

    # Summary sheet: asset type rows, workspace columns
    ws_summary = wb.create_sheet(title="Summary", index=0)
    ws_summary.append(["Asset Type"] + workspace_names)
    for kind in all_kinds_sorted:
        row = [kind] + [counts_per_ws[ws].get(kind, 0) for ws in workspace_names]
        ws_summary.append(row)
    ws_summary.append(["TOTAL"] + [sum(counts_per_ws[ws].values()) for ws in workspace_names])

    # Per-workspace detail sheets
    used_titles: Dict[str, int] = {}
    for ws_name, findings in workspace_results.items():
        base_title = ws_name[:31]
        if base_title in used_titles:
            used_titles[base_title] += 1
            # Reserve space for the numeric suffix (e.g. "_2")
            suffix = f"_{used_titles[base_title]}"
            sheet_title = base_title[: 31 - len(suffix)] + suffix
        else:
            used_titles[base_title] = 1
            sheet_title = base_title
        ws = wb.create_sheet(title=sheet_title)
        ws.append(["kind", "path", "notes", "lockin_count", "lockin_details"])
        for f in sorted(findings, key=lambda x: (x.kind, x.path)):
            ws.append([f.kind, f.path, f.notes, f.lockin_count, f.lockin_details])

    wb.save(out_path)
