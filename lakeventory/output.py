"""Output formatting for inventory results."""

from pathlib import Path
from typing import Dict, List, Tuple

from .config import KIND_TO_SHEET, SHEET_ORDER
from .models import Finding


def summarize(findings: List[Finding], warnings: List[str]) -> Tuple[str, List[str]]:
    """Generate summary and detailed findings.
    
    Args:
        findings: List of discovered findings
        warnings: List of warnings from collection
        
    Returns:
        Tuple of (summary markdown, detail lines)
    """
    counts = {}
    for f in findings:
        counts[f.kind] = counts.get(f.kind, 0) + 1

    summary_lines = ["# Lakeventory Results", ""]
    summary_lines.append("## Summary")
    summary_lines.append("")
    for kind in sorted(counts.keys()):
        summary_lines.append(f"- {kind}: {counts[kind]}")
    summary_lines.append("")

    detail_lines = ["## Findings", ""]
    for f in sorted(findings, key=lambda x: (x.kind, x.path)):
        detail_lines.append(
            f"- [{f.kind}] {f.path} ({f.notes}) | lockin_count={f.lockin_count} | lockin_details={f.lockin_details or '-'}"
        )

    if warnings:
        detail_lines.append("")
        detail_lines.append("## Warnings")
        detail_lines.append("")
        for warning in warnings:
            detail_lines.append(f"- {warning}")

    return "\n".join(summary_lines), detail_lines


def write_markdown(findings: List[Finding], warnings: List[str], out_path: Path) -> None:
    """Write findings to markdown file.
    
    Args:
        findings: List of discovered findings
        warnings: List of warnings from collection
        out_path: Path to output markdown file
    """
    summary, details = summarize(findings, warnings)
    out_path.write_text(summary + "\n" + "\n".join(details) + "\n", encoding="utf-8")


def write_excel(findings: List[Finding], warnings: List[str], out_path: Path) -> None:
    """Write findings to Excel file with categorized sheets.
    
    Args:
        findings: List of discovered findings
        warnings: List of warnings from collection
        out_path: Path to output Excel file
        
    Raises:
        RuntimeError: If openpyxl is not installed
    """
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError("Missing dependency: openpyxl. Install with `pip install openpyxl`.") from exc

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Create sheets in defined order
    sheets = {}
    for sheet_name in SHEET_ORDER:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["kind", "path", "notes", "lockin_count", "lockin_details"])
        sheets[sheet_name] = ws

    # Populate sheets with findings
    for item in findings:
        sheet_name = KIND_TO_SHEET.get(item.kind, "Workspace Objects")
        sheets[sheet_name].append([item.kind, item.path, item.notes, item.lockin_count, item.lockin_details])

    # Add warnings sheet if needed
    if warnings:
        ws = wb.create_sheet(title="Warnings")
        ws.append(["warning"])
        for warning in warnings:
            ws.append([warning])

    wb.save(out_path)


def write_delta_markdown(
    delta_findings: List[Finding],
    stats: Dict,
    warnings: List[str],
    out_path: Path,
) -> None:
    """Write delta report (changes since last run) to markdown.
    
    Args:
        delta_findings: List of added/modified findings
        stats: Dict with added/removed/unchanged/modified counts
        warnings: List of warnings from collection
        out_path: Path to output markdown file
    """
    lines = ["# Lakeventory Delta Report", ""]
    lines.append("## Changes Summary")
    lines.append("")
    lines.append(f"- ✨ **Added:** {stats.get('added', 0)}")
    lines.append(f"- 🗑️ **Removed:** {stats.get('removed', 0)}")
    lines.append(f"- ♻️ **Modified:** {stats.get('modified', 0)}")
    lines.append(f"- ✅ **Unchanged:** {stats.get('unchanged', 0)}")
    lines.append("")
    
    if delta_findings:
        lines.append("## Changed Items")
        lines.append("")
        for f in sorted(delta_findings, key=lambda x: (x.kind, x.path)):
            lines.append(
                f"- [{f.kind}] {f.path} ({f.notes}) | lockin_count={f.lockin_count} | lockin_details={f.lockin_details or '-'}"
            )
    else:
        lines.append("## No Changes")
        lines.append("All inventory items are identical to previous run.")
    
    if warnings:
        lines.append("")
        lines.append("## Warnings")
        lines.append("")
        for warning in warnings:
            lines.append(f"- {warning}")
    
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_delta_excel(
    delta_findings: List[Finding],
    stats: Dict,
    warnings: List[str],
    out_path: Path,
) -> None:
    """Write delta report to Excel file with summary sheet.
    
    Args:
        delta_findings: List of added/modified findings
        stats: Dict with added/removed/unchanged/modified counts
        warnings: List of warnings from collection
        out_path: Path to output Excel file
    """
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError("Missing dependency: openpyxl. Install with `pip install openpyxl`.") from exc

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Summary sheet
    ws_summary = wb.create_sheet(title="Summary", index=0)
    ws_summary.append(["Metric", "Count"])
    ws_summary.append(["Added", stats.get("added", 0)])
    ws_summary.append(["Removed", stats.get("removed", 0)])
    ws_summary.append(["Modified", stats.get("modified", 0)])
    ws_summary.append(["Unchanged", stats.get("unchanged", 0)])
    
    # Changes sheet (categorized)
    sheets = {}
    for sheet_name in SHEET_ORDER:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["kind", "path", "notes", "lockin_count", "lockin_details"])
        sheets[sheet_name] = ws
    
    for item in delta_findings:
        sheet_name = KIND_TO_SHEET.get(item.kind, "Workspace Objects")
        sheets[sheet_name].append([item.kind, item.path, item.notes, item.lockin_count, item.lockin_details])
    
    # Warnings sheet
    if warnings:
        ws = wb.create_sheet(title="Warnings")
        ws.append(["warning"])
        for warning in warnings:
            ws.append([warning])
    
    wb.save(out_path)

